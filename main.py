from PyQt5 import uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QLineEdit, QPushButton, QVBoxLayout, QWidget
from PyQt5.QtCore import QThread, pyqtSignal
import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from front import Ui_MainWindow

# Worker class to handle the background task
class Worker(QThread):
    progress = pyqtSignal(int)  # Signal to update progress bar
    finished = pyqtSignal(bool)  # Signal to indicate task completion

    def __init__(self, folder_path):
        super().__init__()
        self.folder_path = folder_path

    def run(self):
        # Set up Selenium WebDriver with headless Chrome
        options = Options()
        options.add_argument('--headless')
        service = ChromeService(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

        data = []  # List to store data from HTML files
        files = [f for f in os.listdir(self.folder_path) if f.endswith('.html')]
        total_files = len(files)
        for i, filename in enumerate(files):
            filepath = os.path.join(self.folder_path, filename)
            driver.get(f"file:///{filepath}")

            # Extract text content from the HTML file
            body_text = driver.find_element(By.TAG_NAME, 'body').text
            formatted_text = body_text.replace('. ', '.\n')  # Format text

            data.append({'Filename': filename, 'Content': formatted_text})

            # Update progress bar
            self.progress.emit(int((i + 1) / total_files * 100))

        driver.quit()

        # Save data to Excel if there's any
        if data:
            df = pd.DataFrame(data)
            self.save_to_excel_with_style(df)
            self.finished.emit(True)  # Task finished successfully
        else:
            self.finished.emit(False)  # No data found

    def save_to_excel_with_style(self, df):
        # Create an Excel workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "HTML Data"

        # Define styles for the Excel sheet
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Add header row with styles
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
            sheet.column_dimensions[cell.column_letter].width = max(len(header) + 2, 30)

        # Add data rows with styles
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = wrap_alignment
                cell.border = thin_border

        # Adjust column widths based on content
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # Adjust row heights for better readability
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = wrap_alignment
            sheet.row_dimensions[row[0].row].height = 30

        # Save the workbook to a file
        workbook.save('the_excel.xlsx')

# Main window class for the application
class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.browse_btn.clicked.connect(self.browse_folder)  # Connect browse button
        self.run_Button.clicked.connect(self.process_files)  # Connect run button

    def browse_folder(self):
        # Open dialog to select folder
        folder_path = QFileDialog.getExistingDirectory(self, "Select Directory")
        if folder_path:
            self.label_3.setText(folder_path)  # Display selected folder

    def process_files(self):
        folder_path = self.label_3.text()
        if not folder_path:
            self.label_3.setText("No folder selected.")  # Show message if no folder selected
            return
        self.worker = Worker(folder_path)
        self.worker.progress.connect(self.update_progress)  # Connect progress signal
        self.worker.finished.connect(self.on_finished)  # Connect finished signal
        self.worker.start()  # Start the worker thread

    def update_progress(self, value):
        self.progressBar.setValue(value)  # Update progress bar

    def on_finished(self, success):
        if success:
            self.label_3.setText("Excel sheet created successfully: the_excel.xlsx")
        else:
            self.label_3.setText("No HTML files found in the selected folder.")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_win = MainWindow()
    main_win.show()
    sys.exit(app.exec_())
